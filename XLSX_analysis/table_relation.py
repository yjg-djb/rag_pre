from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import logging

# -------------------------- 日志配置（日志文件保存到logs文件夹） --------------------------
def setup_logger(log_file="logs/excel批量处理日志.log"):
    """配置日志：同时输出到控制台和logs文件夹下的日志文件"""
    logger = logging.getLogger("ExcelBatchProcessor")
    logger.setLevel(logging.INFO)
    
    if logger.handlers:
        return logger
    
    # 确保logs文件夹存在
    log_dir = os.path.dirname(log_file)
    if log_dir and not os.path.exists(log_dir):
        os.makedirs(log_dir, exist_ok=True)
        logger.info(f"已自动创建日志文件夹：{os.path.abspath(log_dir)}")
    
    # 日志格式：时间-级别-模块-消息
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - [%(module)s] - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # 控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    
    # 文件处理器（支持中文，保存到logs文件夹）
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    
    return logger

logger = setup_logger()

# -------------------------- 核心处理函数 --------------------------
def auto_detect_header_row(ws, preview_rows=20):
    """
    自动识别表头行（1-based）
    基于3个特征：非空占比≥50%、重复占比≤30%、文本占比≥70%
    """
    header_candidates = []
    for row_idx in range(1, min(preview_rows + 1, ws.max_row + 1)):
        row_cells = []
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row_idx, column=col).value
            row_cells.append(cell_val)
        row_length = len(row_cells)
        if row_length == 0:
            continue
        
        # 特征1：非空单元格占比
        non_null_count = sum(1 for cell in row_cells if cell is not None and str(cell).strip())
        non_null_ratio = non_null_count / row_length
        
        # 特征2：重复值占比
        seen = set()
        duplicates = set()
        for cell in row_cells:
            cell_str = str(cell).strip() if cell is not None else ""
            if cell_str and cell_str in seen:
                duplicates.add(cell_str)
            seen.add(cell_str)
        duplicate_ratio = len(duplicates) / row_length if row_length > 0 else 0
        
        # 特征3：文本单元格占比
        text_count = sum(1 for cell in row_cells if cell is not None and isinstance(cell, str))
        text_ratio = text_count / row_length if row_length > 0 else 0
        
        # 满足特征则视为表头候选
        if non_null_ratio >= 0.5 and duplicate_ratio <= 0.3 and text_ratio >= 0.7:
            header_candidates.append((row_idx, non_null_ratio))
    
    # 选择非空占比最高的候选行，默认第一行
    if header_candidates:
        header_candidates.sort(key=lambda x: x[1], reverse=True)
        selected_header = header_candidates[0][0]
        logger.info(f"表头识别：候选行{[x[0] for x in header_candidates]}，选中第{selected_header}行")
        return selected_header
    logger.warning("未找到符合条件的表头行，默认使用第1行")
    return 1

def flatten_excel_with_merged_cells(excel_path, output_file):
    """
    合并单元格统一赋值处理：
    - 合并单元格区域内所有单元格统一赋值为主单元格的值
    - 保留原始数据类型（数字、日期、文本等）
    - 保留所有行/列、所有原始信息
    """
    try:
        # 加载Excel（支持所有工作表）
        wb = load_workbook(excel_path, data_only=True, read_only=False)
        all_sheets_data = {}  # 存储所有工作表的完整数据

        logger.info("="*80)
        logger.info(f"开始处理Excel文件：{os.path.basename(excel_path)}")
        logger.info(f"文件路径：{os.path.abspath(excel_path)}")
        logger.info(f"共包含 {len(wb.sheetnames)} 个工作表")
        logger.info("="*80)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"\n📋 正在处理工作表：【{sheet_name}】")

            # 获取完整的行数列数（包含所有空行空列）
            def get_full_range(ws):
                max_row = ws.max_row if hasattr(ws, 'max_row') else 1
                max_col = ws.max_column if hasattr(ws, 'max_column') else 1
                return max(max_row, 1), max(max_col, 1)

            total_rows, total_cols = get_full_range(ws)
            logger.info(f"   工作表范围：{total_rows} 行 × {total_cols} 列")

            # ========== 步骤1：先处理所有合并单元格（包括表头行的合并单元格） ==========
            cell_values = {}  # 存储所有单元格的值（包括合并单元格区域）
            merged_ranges = list(ws.merged_cells.ranges)
            logger.info(f"   检测到合并单元格组数：{len(merged_ranges)}")

            if len(merged_ranges) > 0:
                logger.info(f"   开始处理合并单元格，所有合并区域内的单元格将统一赋值...")

            for idx, merge_range in enumerate(merged_ranges, 1):
                min_r, max_r = merge_range.min_row, merge_range.max_row
                min_c, max_c = merge_range.min_col, merge_range.max_col
                
                # 获取主单元格的原始值（保持原始数据类型）
                master_val = ws.cell(row=min_r, column=min_c).value
                
                # 将主单元格的值赋给合并区域内的所有单元格（包括主单元格本身）
                for r in range(min_r, max_r + 1):
                    for c in range(min_c, max_c + 1):
                        cell_values[(r, c)] = master_val
                
                # 日志输出（文本类型才加引号，数字类型直接显示）
                display_val = f"'{master_val}'" if isinstance(master_val, str) else master_val
                cell_count = (max_r - min_r + 1) * (max_c - min_c + 1)
                logger.info(f"      [{idx}/{len(merged_ranges)}] 合并区域 {get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r} "
                           f"({max_r-min_r+1}行×{max_c-min_c+1}列，共{cell_count}个单元格) → 统一赋值为: {display_val}")

            # ========== 步骤2：自动识别表头行 ==========
            header_row = auto_detect_header_row(ws, preview_rows=10)
            logger.info(f"   自动识别表头行：第 {header_row} 行")

            # ========== 步骤3：提取表头（优先使用合并单元格赋值后的值，确保无空值） ==========
            header = []
            empty_header_count = 0  # 统计空表头数量
            for col in range(1, total_cols + 1):
                # 优先使用合并单元格统一赋值后的值
                if (header_row, col) in cell_values:
                    cell_val = cell_values[(header_row, col)]
                else:
                    cell_val = ws.cell(row=header_row, column=col).value
                
                # 处理空表头：如果表头为空、None或只有空格，自动命名为"列X"
                if cell_val is None or str(cell_val).strip() == "":
                    header_val = f"列{col}"
                    empty_header_count += 1
                else:
                    header_val = str(cell_val).strip()
                
                # 处理重复表头：如果表头已存在，添加后缀"_2"、"_3"等
                original_header_val = header_val
                counter = 2
                while header_val in header:
                    header_val = f"{original_header_val}_{counter}"
                    counter += 1
                
                header.append(header_val)
            
            logger.info(f"   提取表头（共{len(header)}列）：{header}")
            if empty_header_count > 0:
                logger.info(f"   ⚠️  检测到 {empty_header_count} 个空表头，已自动命名为「列X」格式")

            # 提取所有数据（从第1行到最后一行，跳过表头行，保留所有其他行）
            full_data = []
            for row in range(1, total_rows + 1):
                if row == header_row:
                    continue  # 跳过表头行（列名已作为DataFrame表头）
                row_data = []
                for col in range(1, total_cols + 1):
                    # 优先使用合并单元格统一赋值的值
                    if (row, col) in cell_values:
                        cell_val = cell_values[(row, col)]
                    else:
                        # 非合并单元格，直接读取原始值
                        cell_val = ws.cell(row=row, column=col).value
                    
                    # 保持原始数据类型，只对None进行转换
                    if cell_val is None:
                        cell_val = ""
                    
                    row_data.append(cell_val)
                full_data.append(row_data)

            # 生成完整DataFrame
            df_flat = pd.DataFrame(full_data, columns=header)
            
            # ========== 过滤完全为空的列 ==========
            original_cols = len(df_flat.columns)
            # 删除所有值都为空的列
            df_flat = df_flat.loc[:, df_flat.apply(lambda col: col.astype(str).str.strip().ne('').any())]
            filtered_cols = len(df_flat.columns)
            removed_cols = original_cols - filtered_cols
            
            if removed_cols > 0:
                logger.info(f"   已自动删除 {removed_cols} 个完全为空的列")
            
            all_sheets_data[sheet_name] = df_flat

            # 输出当前工作表统计信息
            logger.info(f"   处理完成：{len(df_flat)} 行数据，{len(df_flat.columns)} 列字段")

            # 数据量过大时，日志只输出统计信息（避免日志文件过大）
            if len(df_flat) <= 100:
                logger.info(f"   【{sheet_name}】完整数据预览：")
                logger.info(df_flat.to_string(index=False))
            else:
                logger.info(f"   ⚠️  【{sheet_name}】数据量较大（{len(df_flat)} 行），仅输出统计信息")

        # 保存完整数据到Excel（每个工作表对应一个Sheet）
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet_name, df in all_sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # 美化Excel：列宽自适应
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 3, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        # 最终提示
        wb.close()
        logger.info(f"\n" + "="*80)
        logger.info(f"✅ 【{os.path.basename(excel_path)}】所有工作表处理完成！")
        logger.info(f"📁 输出文件路径：{os.path.abspath(output_file)}")
        logger.info(f"📊 输出统计：{len(all_sheets_data)} 个工作表，合并单元格已统一赋值")
        logger.info("="*80)
        
    except Exception as e:
        logger.error(f"❌ 处理文件【{os.path.basename(excel_path)}】时出错：{str(e)}", exc_info=True)
        raise

if __name__ == "__main__":
    # -------------------------- 批量处理配置 --------------------------
    excel_dir = "exal_data"  # 待处理Excel文件存放目录
    output_dir = "exal_solution_result"  # 处理结果输出目录
    output_suffix = "_扁平化结果"  # 输出文件后缀
    
    # -------------------------- 执行批量处理 --------------------------
    logger.info("="*80)
    logger.info("🚀 启动Excel批量处理程序")
    logger.info(f"📂 待处理文件夹：{os.path.abspath(excel_dir)}")
    logger.info(f"📂 结果输出文件夹：{os.path.abspath(output_dir)}")
    logger.info("="*80)
    
    # 检查待处理文件夹是否存在
    if not os.path.exists(excel_dir):
        logger.error(f"❌ 错误：待处理文件夹【{excel_dir}】不存在！")
        logger.error("请确认文件夹名称和路径是否正确，将Excel文件放入该文件夹后重新运行")
        exit(1)
    
    # 确保输出文件夹存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"✅ 已自动创建结果输出文件夹：{os.path.abspath(output_dir)}")
    
    # 筛选出所有Excel文件
    excel_files = [f for f in os.listdir(excel_dir) if f.endswith(('.xlsx', '.xls'))]
    if not excel_files:
        logger.warning(f"⚠️  文件夹【{excel_dir}】中未找到Excel文件（.xlsx/.xls）")
        logger.warning("请检查文件格式是否正确，确保文件未被隐藏")
        exit(0)
    
    # 遍历处理每个Excel文件
    logger.info(f"✅ 找到 {len(excel_files)} 个Excel文件，开始批量处理...")
    for idx, file_name in enumerate(excel_files, 1):
        logger.info(f"\n===== 处理进度：{idx}/{len(excel_files)} - 文件：{file_name} =====")
        excel_path = os.path.join(excel_dir, file_name)
        output_file_name = f"{os.path.splitext(file_name)[0]}{output_suffix}.xlsx"
        output_path = os.path.join(output_dir, output_file_name)
        # 执行处理
        flatten_excel_with_merged_cells(excel_path, output_path)
    
    # 批量处理完成总结
    logger.info("\n" + "="*80)
    logger.info("🎉 所有Excel文件批量处理完成！")
    logger.info(f"📊 处理统计：共处理 {len(excel_files)} 个文件")
    logger.info(f"📁 结果文件：保存在【{os.path.abspath(output_dir)}】文件夹")
    logger.info(f"📜 日志文件：保存在【{os.path.abspath('logs')}】文件夹")
    logger.info("="*80)